from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from django.views import generic
from django.views.generic.base import TemplateView
from django.utils import timezone
import pandas as pd
from mysite import settings
from xlrd.sheet import ctype_text
import re
from django.views.decorators.csrf import csrf_exempt
from xlrd import open_workbook
import openpyxl
from django.shortcuts import render, redirect
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.pagesizes import letter
from django.core.mail import EmailMessage
from django.core.mail import send_mail
import json

Ruta = 'polls/static/polls/Datos/Inventario.xlsx'
@csrf_exempt
def Result(request):
    if request.method == 'POST':
        print('Hola')
        valor =request.POST.get('Valor')
        id=request.POST.get('ID')
        doc = openpyxl.load_workbook(Ruta,data_only=True)
        hoja = doc['inventario']
        hoja[id] = valor
        doc.save(Ruta)
    archivo_excel = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
    ID = archivo_excel['CLASIF.EQUIP(ID)']
    casual = 2
    listaloop=[]
    for contador in ID:
        listaloop.append(casual)
        casual+=1
    Descripcion =archivo_excel['Descripcion']
    EI=archivo_excel['Existencias Iniciales']
    Entradas=archivo_excel['Entradas']
    Salidas=archivo_excel['Salidas']
    Stock=archivo_excel['Stock']
    return render(request ,'polls/Reporte.html',{'ID':ID,'Descripcion':Descripcion,'EI':EI,'Entradas':Entradas,'Salidas':Salidas,'Stock':Stock,'Indice':listaloop})

def Alotes(request):
    if request.method=='POST':
        Codigo = request.POST['Codigo']
        Fecha = request.POST['Fecha']
        Factura = request.POST['Factura']
        Descripcion = request.POST['Descripcion']
        Cantidad = request.POST['Cantidad']
        Cantidad=int(Cantidad)
        Serie = request.POST['Serie']
        ObservacionesP = request.POST['Observaciones']
        controlador = request.POST['indice']
        print(controlador)
        controlador=json.loads(controlador)
        FacturaP=json.loads(Factura)
        CodigoP=json.loads(Codigo)
        DescripcionP=json.loads(Descripcion)
        SerieP=json.loads(Serie)
        FechaP = json.loads(Fecha)

        # Validacion de sobreescribir en datos vacios Obtenemos los datos que hay en la hoja de entradas y lo almacenamos en variables
        archivo_excel = pd.read_excel(Ruta, sheet_name='Entradas', keep_default_na=False)
        Factura = archivo_excel['Num de Factura']
        Fecha = archivo_excel['Fecha']
        Codigo = archivo_excel['Codigo de Producto']
        Descripcion = archivo_excel['Descripcion']
        CantidadP = archivo_excel['Cantidad']
        Serie = archivo_excel['SERIE']
        Observaciones = archivo_excel['OBSERVACIONES']
        print("Se Realizara una nueva entrada")

        #Validacion de sumas
        archivo_excel2 = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
        CantidadI = archivo_excel2['Existencias Iniciales']
        Entradas = archivo_excel2['Entradas']
        Salidas = archivo_excel2['Salidas']
        doc = openpyxl.load_workbook(Ruta, data_only=True)
        hoja = doc['Entradas']
        Sumador = doc['inventario']

        #Bucle Para el ingreso de cantidades por lotes
        V=0
        for Go in controlador:
            controlador[V] = int(controlador[V])
            SumaV = []
            SumaV.append(CantidadI[controlador[V]])
            SumaV.append(Entradas[controlador[V]])
            SumaV.append(Salidas[controlador[V]])
            print(SumaV)
            Totalstock = ((SumaV[0] + SumaV[1] + Cantidad) - (SumaV[2]))
            TotalS = SumaV[1] + Cantidad
            indicador = controlador[V] + 2
            indicador = str(indicador)
            #Aqui se empiezan a agregar datos
            Sumador['D'+ indicador ] = TotalS
            Sumador['F'+ indicador] = Totalstock
            V += 1
            doc.save(Ruta)
            archivo_excel2 = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
            CantidadI = archivo_excel2['Existencias Iniciales']
            Entradas = archivo_excel2['Entradas']
            Salidas = archivo_excel2['Salidas']
            doc = openpyxl.load_workbook(Ruta, data_only=True)
            hoja = doc['Entradas']
            Sumador = doc['inventario']

        #En esta parte se realizara la entrada lamentablemente se realizan muchisimas entradas y se tiene que validar
        #La entrada de cada uno de los productos ya uqe es por lotes
        V=0
        for X in CodigoP:
            i = 0
            validador = 1
            indice = 0
            if Codigo.empty == True and Descripcion.empty == True and CantidadP.empty == True and Factura.empty == True and Fecha.empty == True and Serie.empty == True and Observaciones.empty == True:
                indice = 2
                validador = 0
                i = 0
            else:
                for Rol in Codigo:
                    if str(Codigo[i]).strip(" ") == "" and str(Descripcion[i]).strip(" ") == "" and str(CantidadP[i]).strip(" ") == "":
                        indice = i + 2
                        validador = 0
                        break
                    else:
                        indice = i + 3
                        validador = 0
                    i += 1
            indice = str(indice)
            if validador == 1:#Deabilite esta columna para ver si este es el error Todas las que tengan * no se estaran utilizando
                hoja.append([Factura[V],Fecha[V],CodigoP[V],Descripcion[V],Cantidad,Serie[V],ObservacionesP])#*
            else:
                hoja['A' + indice] =FacturaP[V]
                hoja['B' + indice] =FechaP[V]
                hoja['C' + indice] =CodigoP[V]
                hoja['D' + indice] =DescripcionP[V]
                hoja['E' + indice] =Cantidad
                hoja['F' + indice] =SerieP[V]
                hoja['G' + indice] =ObservacionesP
            doc.save(Ruta)
            archivo_excel = pd.read_excel(Ruta, sheet_name='Entradas', keep_default_na=False)
            Factura = archivo_excel['Num de Factura']
            Fecha = archivo_excel['Fecha']
            Codigo = archivo_excel['Codigo de Producto']
            Descripcion = archivo_excel['Descripcion']
            CantidadP = archivo_excel['Cantidad']
            Serie = archivo_excel['SERIE']
            Observaciones = archivo_excel['OBSERVACIONES']
            V+=1
        return render(request,'polls/Alotes.html')

    else:
        archivo_excel = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
        Codigo = archivo_excel['CLASIF.EQUIP(ID)']
        CodigoP = []
        Descripcion = archivo_excel['Descripcion']
        DescripcionP = []
        for vuelta in Descripcion:
            if str(vuelta).strip(' ') != "":
                DescripcionP.append(vuelta)
            else:
                DescripcionP.append('Sin Descripcion')
        for vuelta in Codigo:
            if str(vuelta).strip(' ') != "":
                CodigoP.append(vuelta)
            else:
                CodigoP.append('Sin Codigo')
        return render(request, 'polls/Alotes.html', {'Codigo': CodigoP, 'Descripcion': DescripcionP})

def Ninventario(request):
    if request.method == 'POST':
        # Validacion de sobreescribir en datos vacios
        archivo_excel = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
        Codigo = archivo_excel['CLASIF.EQUIP(ID)']
        Descripcion = archivo_excel['Descripcion']
        CantidadP = archivo_excel['Existencias Iniciales']
        Entradas=archivo_excel['Entradas']
        Salidas=archivo_excel['Salidas']
        Stock=archivo_excel['Stock']
        i = 0
        validador=1
        indice=0
        if Codigo.empty==True and Descripcion.empty==True and CantidadP.empty==True and Entradas.empty==True and Salidas.empty==True and Stock.empty==True:
            indice=2
            validador=0
        else:
            for Rol in Codigo:
                if str(Codigo[i]).strip(' ') == "" and str(Descripcion[i]).strip(' ') == "" and str(CantidadP[i]).strip(' ') == "" and str(Entradas[i]).strip(' ')=='' and str(Salidas[i]).strip(' ')=='' and str(Stock[i]).strip(' ') =='' :
                    indice=i+2
                    validador=0
                    print(indice)
                    break
                else:
                    indice = i + 3
                    validador=0
                i +=1
        indice = str(indice)
        print("Se valido todas celdas vacias")
        IDI = request.POST['Identificador']
        Descripcion = request.POST['Descripcion']
        CantidadI = request.POST['ExistenciaI']
        ret = int(CantidadI)
        doc = openpyxl.load_workbook(Ruta, data_only=True)

        hoja = doc['inventario']
        print(IDI)
        if validador == 1:#*
            hoja.append([IDI,Descripcion,ret,0,0,ret])#*
        else:
            hoja['A' + indice] = IDI
            hoja['B' + indice] = Descripcion
            hoja['C' + indice] = ret
            hoja['D' + indice] = 0
            hoja['E' + indice] = 0
            hoja['F' + indice] = ret

        doc.save(Ruta)
        return HttpResponseRedirect(reverse('polls:Reporte'))
    else:
        return render(request, 'polls/Ninventario.html')

def Entradas(request):
    archivo_excel = pd.read_excel(Ruta, sheet_name='Entradas', keep_default_na=False)
    Factura=archivo_excel['Num de Factura']
    casual = 2
    listaloop=[]
    for contador in Factura:
        listaloop.append(casual)
        casual+=1
    Fecha=archivo_excel['Fecha']
    Codigo=archivo_excel['Codigo de Producto']
    Descripcion=archivo_excel['Descripcion']
    Cantidad=archivo_excel['Cantidad']
    Serie=archivo_excel['SERIE']
    Observaciones=archivo_excel['OBSERVACIONES']
    return render(request ,'polls/Entradas.html',{'Factura':Factura,'Fecha':Fecha,'Codigo':Codigo,'Descripcion':Descripcion,'Cantidad':Cantidad,'Serie':Serie,'Observaciones':Observaciones,'Indice':listaloop})

def Nentrada(request):
    if request.method == 'POST':
        # Validacion de sobreescribir en datos vacios
        archivo_excel = pd.read_excel(Ruta, sheet_name='Entradas', keep_default_na=False)
        Factura=archivo_excel['Num de Factura']
        Fecha=archivo_excel['Fecha']
        Codigo = archivo_excel['Codigo de Producto']
        Descripcion = archivo_excel['Descripcion']
        CantidadP = archivo_excel['Cantidad']
        Serie=archivo_excel['SERIE']
        Observaciones=archivo_excel['OBSERVACIONES']
        i = 0
        validador = 1
        indice = 0
        if Codigo.empty==True and Descripcion.empty==True and CantidadP.empty==True and Factura.empty==True and Fecha.empty==True and Serie.empty==True and Observaciones.empty==True:
            indice=2
            validador=0
            i=0
        else:
            for Rol in Codigo:
                print(Codigo[i])
                print(Descripcion[i])
                if str(Codigo[i]).strip(" ") == "" and str(Descripcion[i]).strip(" ") == "" and str(CantidadP[i]).strip(" ") == "":
                    indice = i + 2
                    validador = 0
                    break
                else:
                    indice = i + 3
                    validador=0
                i += 1
        indice = str(indice)
        print("Se Realizara una nueva entrada")
        Factura = request.POST['Factura']
        Fecha = request.POST['Fecha']
        CodigoP = request.POST['CodigoP']
        Descripcion=request.POST['Descripcion']
        Cantidad=request.POST['Cantidad']
        ret = int(Cantidad)
        Serie=request.POST['SERIE']
        Observaciones=request.POST['OBSERVACIONES']
        #Validacion de sumas
        archivo_excel2 = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
        CantidadI = archivo_excel2['Existencias Iniciales']
        Entradas = archivo_excel2['Entradas']
        Salidas = archivo_excel2['Salidas']
        Controlador = request.POST['Controlador']
        Controlador = int(Controlador)
        SumaV = []
        SumaV.append(CantidadI[Controlador])
        SumaV.append(Entradas[Controlador])
        SumaV.append(Salidas[Controlador])
        Totalstock = (SumaV[0] + SumaV[1] + ret) - (SumaV[2])
        TotalS = SumaV[1] + ret
        indicador = Controlador + 2
        indicador = str(indicador)

        #Aqui se empiezan a agregar datos
        doc = openpyxl.load_workbook(Ruta, data_only=True)
        hoja = doc['Entradas']
        Sumador=doc['inventario']
        Sumador['D'+ indicador ] = TotalS
        Sumador['F'+ indicador] = Totalstock
        if validador == 1:#Deabilite esta columna para ver si este es el error Todas las que tengan * no se estaran utilizando
            hoja.append([Factura,Fecha,CodigoP,Descripcion,ret,Serie,Observaciones])#*
        else:
            hoja['A' + indice] = Factura
            hoja['B' + indice] = Fecha
            hoja['C' + indice] = CodigoP
            hoja['D' + indice] = Descripcion
            hoja['E' + indice] = ret
            hoja['F' + indice] = Serie
            hoja['G' + indice] = Observaciones
        doc.save(Ruta)
        return HttpResponseRedirect(reverse('polls:Entradas'))
    else:
        archivo_excel = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
        Codigo = archivo_excel['CLASIF.EQUIP(ID)']
        CodigoP=[]
        Descripcion = archivo_excel['Descripcion']
        DescripcionP=[]
        for vuelta in Descripcion:
            if str(vuelta).strip(' ') != "":
                DescripcionP.append(vuelta)
            else:
                DescripcionP.append('Sin Descripcion')
        for vuelta in Codigo:
            if str(vuelta).strip(' ') != "":
                CodigoP.append(vuelta)
            else:
                CodigoP.append('Sin Codigo')
        return render(request, 'polls/Nentrada.html',{'Codigo':CodigoP,'Descripcion':DescripcionP})

def Salidas(request):
    archivo_excel = pd.read_excel(Ruta, sheet_name='Salidas', keep_default_na=False)
    Factura=archivo_excel['Num de Factura']
    casual = 2
    listaloop=[]
    for contador in Factura:
        listaloop.append(casual)
        casual+=1
    Fecha=archivo_excel['Fecha']
    Codigo=archivo_excel['Codigo Producto']
    Descripcion=archivo_excel['Descripcion(Producto)']
    Cantidad=archivo_excel['Cantidad']
    Serie=archivo_excel['Serie']
    Observaciones=archivo_excel['OBSERVACIONES']
    Cliente=archivo_excel['DE O P/CLIENTE']
    Recibio=archivo_excel['TECNICO RECIBIO Y FIRMO SALIDA']
    return render(request ,'polls/Salidas.html',{'Cliente':Cliente,'Recibio':Recibio,'Factura':Factura,'Fecha':Fecha,'Codigo':Codigo,'Descripcion':Descripcion,'Cantidad':Cantidad,'Serie':Serie,'Observaciones':Observaciones,'Indice':listaloop})

def Nsalida(request):
    if request.method == 'POST':
        #Se Realiza una validacion para buscar datos en blanco y sustituirlos
        archivo_excel = pd.read_excel(Ruta, sheet_name='Salidas', keep_default_na=False)
        Factura=archivo_excel['Num de Factura']
        Fecha=archivo_excel['Fecha']
        Codigo = archivo_excel['Codigo Producto']
        Descripcion = archivo_excel['Descripcion(Producto)']
        CantidadP = archivo_excel['Cantidad']
        OBSERVACIONES =archivo_excel['OBSERVACIONES']
        Cliente=archivo_excel['DE O P/CLIENTE']
        Tecnico=archivo_excel['TECNICO RECIBIO Y FIRMO SALIDA']
        Serie=archivo_excel['Serie']
        i = 0
        validador=1
        indice=0
        if Factura.empty == True and Fecha.empty==True and Codigo.empty==True and Descripcion.empty==True and CantidadP.empty==True and OBSERVACIONES.empty==True and Cliente.empty==True and Tecnico.empty ==True and Serie.empty ==True:
            indice=2
            validador=0
        else:
            for Rol in Codigo:
                if str(Factura[i]).strip(' ')=='' and str(Fecha[i]).strip(' ')=='' and str(Codigo[i]).strip(" ") == "" and str(Descripcion[i]).strip(" ") == "" and str(CantidadP[i]).strip(" ") == "" and str(OBSERVACIONES[i]).strip(' ')=='' and str(Cliente[i]).strip(' ')=='' and str(Tecnico[i]).strip(' ')=='' and str(Serie[i]).strip(' ')=='':
                    indice=i+2
                    validador=0
                    break
                else:
                    indice = i + 3
                    validador=0
                i +=1
        indice = str(indice)
        print(indice)
        #Se procede a obtener los datos del form para agregar en la tabla salidas
        print("Se agregara una nueva salida")
        Factura = request.POST['Factura']
        Fecha = request.POST['Fecha']
        CodigoP = request.POST['CodigoP']
        Descripcion=request.POST['Descripcion']
        Cantidad=request.POST['Cantidad']
        ret = int(Cantidad)
        Observaciones=request.POST['OBSERVACIONES']
        Cliente=request.POST['Cliente']
        Tecnico=request.POST['Tecnico']
        Serie = request.POST['SERIE']
        RM=request.POST['Respuesta']

        #Aqui se realiza la formula para ingresar las salidas en el inventario con su respectivo valor (Obtener el indice)
        archivo_excel2 = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
        CantidadI=archivo_excel2['Existencias Iniciales']
        Entradas=archivo_excel2['Entradas']
        Salidas=archivo_excel2['Salidas']
        Controlador = request.POST['Controlador']
        Controlador =int(Controlador)
        SumaV=[]
        SumaV.append(CantidadI[Controlador])
        SumaV.append(Entradas[Controlador])
        SumaV.append(Salidas[Controlador])
        Totalstock =(SumaV[0]+SumaV[1]) - (SumaV[2]+ret)
        TotalS=SumaV[2]+ret
        indicador = Controlador + 2
        indicador = str(indicador)

        #Se procede a agregar los datos despues de todas las validaciones
        doc = openpyxl.load_workbook(Ruta, data_only=True)
        hoja = doc['Salidas']
        Sumador=doc['inventario']
        Sumador['E'+ indicador ] = TotalS
        Sumador['F'+ indicador] = Totalstock
        if validador==1:#*
            hoja.append([Factura,Fecha,CodigoP,Descripcion,ret,Observaciones,Cliente,Tecnico,Serie])#*
        else:
            hoja['A' + indice] = Factura
            hoja['B' + indice] = Fecha
            hoja['C' + indice] = CodigoP
            hoja['D' + indice] = Descripcion
            hoja['E' + indice] = ret
            hoja['F' + indice] = Observaciones
            hoja['G' + indice] = Cliente
            hoja['H' + indice] = Tecnico
            hoja['I' + indice] = Serie
        doc.save(Ruta)
        if RM =="Si":
            Reporteador('hola','hola','hola','hola','Hola')
            EnviarMail('hola')
        return HttpResponseRedirect(reverse('polls:Salidas'))
    else:

        #Con este fregmento de codigo obtenemos los datos de los codigos y descripciones que estan en el inventario para automatizar el ingreso de datos
        archivo_excel = pd.read_excel(Ruta, sheet_name='inventario', keep_default_na=False)
        Codigo = archivo_excel['CLASIF.EQUIP(ID)']
        Descripcion = archivo_excel['Descripcion']
        CodigoP=[]
        DescripcionP=[]
        for vuelta in Descripcion:

            if str(vuelta).strip(" ") != "":
                DescripcionP.append(vuelta)
            else:
                DescripcionP.append('Sin descripcion')
        for vuelta in Codigo:
            if str(vuelta).strip(" ") != "":
                CodigoP.append(vuelta)
            else:
                CodigoP.append('Sin codigo')
        return render(request, 'polls/Nsalida.html',{'Codigo':CodigoP,'Descripcion':DescripcionP})

def xls_KeyValue(file, row_KeySeacrh, col_KeySearch):
    pos_RowSearch = None
    pos_ColSearch = None
    wb = open_workbook(file, 'rb')
    sh = wb.sheet_by_index(0)
    row_KeySeacrh = re.compile(row_KeySeacrh)
    col_KeySearch = re.compile(col_KeySearch)

    for rows in range(sh.nrows):
        rows_int = rows  # Stores the number of row where row_KeySearch will match
        rows = str(sh.row(rows))

        if (row_KeySeacrh.search(rows)):
            pos_RowSearch = rows_int
            for cols in range(sh.ncols):
                cols_int = cols  # Stores the number of col
                cols = str(sh.col(cols))
                if col_KeySearch.search(cols):
                    pos_ColSearch = cols_int
                    print('El numero de la fila es :' ,+ cols_int)
                    return ((sh.cell(pos_RowSearch, pos_ColSearch)))
                else:
                    continue
        else:
            continue
Buscador = xls_KeyValue(Ruta, 'hola crayola', 'Descripcion')
print(Buscador)

def Login(request):
    return render(request, 'polls/Login.html')

def EnviarMail(Nombre):
    email = EmailMessage('Reporte', 'Esto es una prueba', to=['samuelsolano509@gmail.com'])
    email.attach_file("polls/static/polls/Datos/Reportes/"+Nombre+".pdf")
    email.send()

def Reporteador(Nombre,TituloR,Fecha,Metodo,Codigo):
    c = canvas.Canvas("polls/static/polls/Datos/Reportes/"+Nombre+".pdf",pagesize = A4 )
    c.setLineWidth(.3)
    c.setFont('Helvetica', 12)

    c.drawCentredString(A4[0] / 2, 810, u"Reporte de "+Metodo)
    c.drawString(30, 750, TituloR)
    c.drawString(30, 735, 'Intelvid Telecom')
    c.drawString(470, 750, Fecha)
    c.line(470, 747, 580, 747)

    c.drawString(325, 725, 'ESTIMADO:')
    c.drawString(415, 725, "Gestor De Entradas y Salidas")
    c.line(398, 723, 580, 723)

    c.drawString(30, 703, 'ASUNTO:')
    c.line(120, 700, 580, 700)
    c.drawString(120, 703, "Reporte De Una Entrada O Salida")

    c.drawString(50,600,"Por este medio se hace constar de la realizacion de una "+Metodo+" La cual fue realizada")
    c.drawString(50,580,"En la Fecha "+Fecha)

    c.save()
